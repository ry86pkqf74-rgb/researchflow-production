"""
Sheet Reader Module - Streaming CSV/Excel reader for large files.

This module provides memory-efficient reading of large spreadsheet files
(CSV, XLSX) with support for chunked processing and optional Dask integration.

Key Features:
- Automatic file type detection
- Streaming CSV with configurable chunk size
- Read-only Excel streaming via openpyxl
- Optional Dask parallel scanning
- Memory usage monitoring

Architecture Notes:
- Returns generators for memory efficiency
- Chunk size configurable (default 50K rows)
- Excel sheets read in read_only mode
"""

import os
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import (
    Iterator, Optional, List, Dict, Any, 
    Union, Generator, Tuple, Literal
)
from datetime import datetime, timezone
import pandas as pd

logger = logging.getLogger(__name__)

# Optional imports with availability flags
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel support disabled")

try:
    import dask.dataframe as dd
    DASK_AVAILABLE = True
except ImportError:
    DASK_AVAILABLE = False
    logger.info("dask not available - parallel processing disabled")


@dataclass
class SheetMetadata:
    """Metadata about a spreadsheet file."""
    path: str
    file_type: Literal["csv", "xlsx", "xls", "tsv"]
    file_size_bytes: int
    file_size_mb: float
    estimated_rows: Optional[int] = None
    column_count: Optional[int] = None
    columns: List[str] = field(default_factory=list)
    sheet_names: List[str] = field(default_factory=list)  # For Excel
    detected_encoding: Optional[str] = None
    scan_timestamp: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "path": self.path,
            "file_type": self.file_type,
            "file_size_bytes": self.file_size_bytes,
            "file_size_mb": self.file_size_mb,
            "estimated_rows": self.estimated_rows,
            "column_count": self.column_count,
            "columns": self.columns,
            "sheet_names": self.sheet_names,
            "detected_encoding": self.detected_encoding,
            "scan_timestamp": self.scan_timestamp,
        }


@dataclass
class ChunkResult:
    """Result of reading a single chunk."""
    chunk_index: int
    row_start: int
    row_end: int
    row_count: int
    df: pd.DataFrame
    columns: List[str]
    memory_bytes: int
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "chunk_index": self.chunk_index,
            "row_start": self.row_start,
            "row_end": self.row_end,
            "row_count": self.row_count,
            "columns": self.columns,
            "memory_bytes": self.memory_bytes,
        }


def detect_file_type(path: Union[str, Path]) -> Literal["csv", "xlsx", "xls", "tsv"]:
    """
    Detect spreadsheet file type from extension.
    
    Args:
        path: Path to file
        
    Returns:
        File type string
        
    Raises:
        ValueError: If file type not supported
    """
    path = Path(path)
    ext = path.suffix.lower()
    
    type_map = {
        ".csv": "csv",
        ".xlsx": "xlsx",
        ".xls": "xls",
        ".tsv": "tsv",
    }
    
    if ext not in type_map:
        raise ValueError(f"Unsupported file type: {ext}. Supported: {list(type_map.keys())}")
    
    return type_map[ext]


def get_sheet_metadata(
    path: Union[str, Path],
    sample_rows: int = 1000,
) -> SheetMetadata:
    """
    Get metadata about a spreadsheet file without loading it fully.
    
    Args:
        path: Path to spreadsheet file
        sample_rows: Number of rows to sample for estimation
        
    Returns:
        SheetMetadata object
    """
    path = Path(path)
    
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    
    file_size = path.stat().st_size
    file_type = detect_file_type(path)
    
    metadata = SheetMetadata(
        path=str(path),
        file_type=file_type,
        file_size_bytes=file_size,
        file_size_mb=file_size / (1024 * 1024),
    )
    
    # Get columns and estimate rows
    if file_type in ("csv", "tsv"):
        sep = "\t" if file_type == "tsv" else ","
        
        # Read header and sample
        try:
            df_sample = pd.read_csv(path, sep=sep, nrows=sample_rows)
            metadata.columns = list(df_sample.columns)
            metadata.column_count = len(df_sample.columns)
            
            # Estimate total rows from file size
            if len(df_sample) > 0:
                bytes_per_row = file_size / len(df_sample) if len(df_sample) >= sample_rows else None
                if bytes_per_row:
                    metadata.estimated_rows = int(file_size / bytes_per_row)
                else:
                    # File smaller than sample - exact count
                    metadata.estimated_rows = len(df_sample)
                    
        except Exception as e:
            logger.warning(f"Failed to sample CSV: {e}")
            
    elif file_type in ("xlsx", "xls"):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel files")
        
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            metadata.sheet_names = wb.sheetnames
            
            # Get first sheet info
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
                metadata.columns = [
                    str(cell.value) if cell.value else f"col_{i}"
                    for i, cell in enumerate(next(ws.iter_rows(max_row=1)))
                ]
                metadata.column_count = len(metadata.columns)
                metadata.estimated_rows = ws.max_row - 1 if ws.max_row else 0
            
            wb.close()
            
        except Exception as e:
            logger.warning(f"Failed to read Excel metadata: {e}")
    
    return metadata


def read_csv_chunks(
    path: Union[str, Path],
    chunk_rows: int = 50_000,
    columns: Optional[List[str]] = None,
    dtype: Optional[Dict[str, Any]] = None,
    encoding: str = "utf-8",
) -> Generator[ChunkResult, None, None]:
    """
    Read a CSV file in chunks.
    
    Args:
        path: Path to CSV file
        chunk_rows: Number of rows per chunk
        columns: Specific columns to read (None = all)
        dtype: Column dtypes
        encoding: File encoding
        
    Yields:
        ChunkResult for each chunk
    """
    path = Path(path)
    file_type = detect_file_type(path)
    sep = "\t" if file_type == "tsv" else ","
    
    reader = pd.read_csv(
        path,
        sep=sep,
        chunksize=chunk_rows,
        usecols=columns,
        dtype=dtype,
        encoding=encoding,
        on_bad_lines="warn",
    )
    
    row_offset = 0
    for chunk_idx, df in enumerate(reader):
        yield ChunkResult(
            chunk_index=chunk_idx,
            row_start=row_offset,
            row_end=row_offset + len(df) - 1,
            row_count=len(df),
            df=df,
            columns=list(df.columns),
            memory_bytes=df.memory_usage(deep=True).sum(),
        )
        row_offset += len(df)


def read_excel_chunks(
    path: Union[str, Path],
    sheet_name: Optional[str] = None,
    chunk_rows: int = 50_000,
    columns: Optional[List[str]] = None,
) -> Generator[ChunkResult, None, None]:
    """
    Read an Excel file in chunks using streaming mode.
    
    Args:
        path: Path to Excel file
        sheet_name: Sheet to read (None = first sheet)
        chunk_rows: Number of rows per chunk
        columns: Specific columns to read (None = all)
        
    Yields:
        ChunkResult for each chunk
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required for Excel files")
    
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    
    try:
        # Select sheet
        ws_name = sheet_name or wb.sheetnames[0]
        ws = wb[ws_name]
        
        # Get headers from first row
        headers = None
        rows_buffer = []
        chunk_idx = 0
        row_offset = 0
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(cell) if cell else f"col_{i}" for i, cell in enumerate(row)]
                continue
            
            # Filter columns if specified
            if columns:
                col_indices = [headers.index(c) for c in columns if c in headers]
                row_values = [row[i] if i < len(row) else None for i in col_indices]
                use_headers = [headers[i] for i in col_indices]
            else:
                row_values = list(row)
                use_headers = headers
            
            rows_buffer.append(row_values)
            
            # Yield chunk when buffer full
            if len(rows_buffer) >= chunk_rows:
                df = pd.DataFrame(rows_buffer, columns=use_headers)
                yield ChunkResult(
                    chunk_index=chunk_idx,
                    row_start=row_offset,
                    row_end=row_offset + len(df) - 1,
                    row_count=len(df),
                    df=df,
                    columns=use_headers,
                    memory_bytes=df.memory_usage(deep=True).sum(),
                )
                chunk_idx += 1
                row_offset += len(df)
                rows_buffer = []
        
        # Yield remaining rows
        if rows_buffer:
            df = pd.DataFrame(rows_buffer, columns=use_headers if use_headers else headers)
            yield ChunkResult(
                chunk_index=chunk_idx,
                row_start=row_offset,
                row_end=row_offset + len(df) - 1,
                row_count=len(df),
                df=df,
                columns=list(df.columns),
                memory_bytes=df.memory_usage(deep=True).sum(),
            )
            
    finally:
        wb.close()


def read_sheet_chunks(
    path: Union[str, Path],
    chunk_rows: int = 50_000,
    sheet_name: Optional[str] = None,
    columns: Optional[List[str]] = None,
    dtype: Optional[Dict[str, Any]] = None,
) -> Generator[ChunkResult, None, None]:
    """
    Read any supported spreadsheet file in chunks.
    
    Auto-detects file type and uses appropriate reader.
    
    Args:
        path: Path to spreadsheet file
        chunk_rows: Number of rows per chunk
        sheet_name: Sheet name for Excel files
        columns: Specific columns to read
        dtype: Column dtypes (CSV only)
        
    Yields:
        ChunkResult for each chunk
    """
    path = Path(path)
    file_type = detect_file_type(path)
    
    if file_type in ("csv", "tsv"):
        yield from read_csv_chunks(
            path=path,
            chunk_rows=chunk_rows,
            columns=columns,
            dtype=dtype,
        )
    elif file_type in ("xlsx", "xls"):
        yield from read_excel_chunks(
            path=path,
            sheet_name=sheet_name,
            chunk_rows=chunk_rows,
            columns=columns,
        )
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


def read_with_dask(
    path: Union[str, Path],
    blocksize: str = "64MB",
    columns: Optional[List[str]] = None,
) -> "dd.DataFrame":
    """
    Read CSV file using Dask for parallel processing.
    
    Args:
        path: Path to CSV file
        blocksize: Dask block size
        columns: Specific columns to read
        
    Returns:
        Dask DataFrame
        
    Raises:
        ImportError: If dask not available
    """
    if not DASK_AVAILABLE:
        raise ImportError("dask required for parallel processing")
    
    path = Path(path)
    file_type = detect_file_type(path)
    
    if file_type not in ("csv", "tsv"):
        raise ValueError("Dask parallel reading only supports CSV/TSV files")
    
    sep = "\t" if file_type == "tsv" else ","
    
    ddf = dd.read_csv(
        str(path),
        sep=sep,
        blocksize=blocksize,
        usecols=columns,
        assume_missing=True,
    )
    
    return ddf


class SheetReader:
    """
    High-level sheet reader with configuration support.
    
    Example:
        reader = SheetReader(chunk_rows=50000)
        
        # Get metadata
        metadata = reader.get_metadata("data.csv")
        
        # Stream chunks
        for chunk in reader.read_chunks("data.csv"):
            process(chunk.df)
    """
    
    def __init__(
        self,
        chunk_rows: int = 50_000,
        enable_dask: bool = False,
        dask_blocksize: str = "64MB",
    ):
        """
        Initialize sheet reader.
        
        Args:
            chunk_rows: Default chunk size
            enable_dask: Use Dask for parallel processing
            dask_blocksize: Dask block size
        """
        self.chunk_rows = chunk_rows
        self.enable_dask = enable_dask and DASK_AVAILABLE
        self.dask_blocksize = dask_blocksize
    
    def get_metadata(self, path: Union[str, Path]) -> SheetMetadata:
        """Get metadata for a spreadsheet file."""
        return get_sheet_metadata(path)
    
    def read_chunks(
        self,
        path: Union[str, Path],
        sheet_name: Optional[str] = None,
        columns: Optional[List[str]] = None,
    ) -> Generator[ChunkResult, None, None]:
        """
        Read spreadsheet in chunks.
        
        Args:
            path: Path to file
            sheet_name: Sheet name (Excel only)
            columns: Columns to read
            
        Yields:
            ChunkResult for each chunk
        """
        yield from read_sheet_chunks(
            path=path,
            chunk_rows=self.chunk_rows,
            sheet_name=sheet_name,
            columns=columns,
        )
    
    def read_dask(
        self,
        path: Union[str, Path],
        columns: Optional[List[str]] = None,
    ) -> "dd.DataFrame":
        """
        Read CSV with Dask (parallel processing).
        
        Args:
            path: Path to CSV file
            columns: Columns to read
            
        Returns:
            Dask DataFrame
        """
        if not self.enable_dask:
            raise RuntimeError("Dask not enabled for this reader")
        
        return read_with_dask(
            path=path,
            blocksize=self.dask_blocksize,
            columns=columns,
        )
    
    def is_large_file(self, path: Union[str, Path], threshold_mb: float = 200) -> bool:
        """Check if file exceeds size threshold."""
        metadata = self.get_metadata(path)
        return metadata.file_size_mb > threshold_mb


__all__ = [
    "SheetMetadata",
    "ChunkResult",
    "detect_file_type",
    "get_sheet_metadata",
    "read_csv_chunks",
    "read_excel_chunks",
    "read_sheet_chunks",
    "read_with_dask",
    "SheetReader",
    "OPENPYXL_AVAILABLE",
    "DASK_AVAILABLE",
]
