"""
Literature Review Matrix Builder

Generates structured literature review matrices in CSV/Excel format.
Columns: Title, Year, DOI/PMID, Population, Design, Cohort Size, Outcomes, Key Findings, Notes, Relevance Score
"""

from __future__ import annotations

import os
import json
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional
from pathlib import Path
from datetime import datetime

from src.governance.output_phi_guard import guard_text

logger = logging.getLogger(__name__)

ARTIFACTS_PATH = os.getenv("ARTIFACTS_PATH", "/data/artifacts")


@dataclass
class MatrixRow:
    """Single row in literature matrix"""
    title: str
    year: Optional[int]
    identifier: str  # DOI or PMID
    population: str
    design: str
    cohort_size: str
    outcomes: str
    key_findings: str
    notes: str
    relevance_score: float


@dataclass
class LiteratureMatrix:
    """Complete literature review matrix"""
    rows: List[MatrixRow]
    query: str
    generated_at: str
    paper_count: int
    csv_path: Optional[str] = None
    xlsx_path: Optional[str] = None


def _extract_population(abstract: str) -> str:
    """Extract population information from abstract"""
    population_terms = [
        'patients', 'participants', 'subjects', 'adults', 'children',
        'elderly', 'women', 'men', 'pediatric', 'adolescents'
    ]

    abstract_lower = abstract.lower()

    for term in population_terms:
        if term in abstract_lower:
            # Find the context around the term
            idx = abstract_lower.find(term)
            start = max(0, idx - 30)
            end = min(len(abstract), idx + len(term) + 30)
            context = abstract[start:end].strip()

            # Clean up
            if not context.startswith(('a', 'A', 't', 'T', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9')):
                context = '...' + context
            if not context.endswith(('.', '!', '?')):
                context = context + '...'

            return context

    return "Not specified"


def _extract_study_design(abstract: str) -> str:
    """Extract study design from abstract"""
    designs = {
        'randomized controlled trial': 'RCT',
        'rct': 'RCT',
        'randomized': 'RCT',
        'cohort study': 'Cohort',
        'cohort': 'Cohort',
        'prospective': 'Prospective',
        'retrospective': 'Retrospective',
        'case-control': 'Case-Control',
        'cross-sectional': 'Cross-Sectional',
        'systematic review': 'Systematic Review',
        'meta-analysis': 'Meta-Analysis',
        'observational': 'Observational',
        'longitudinal': 'Longitudinal',
        'case study': 'Case Study',
        'case report': 'Case Report',
        'survey': 'Survey',
    }

    abstract_lower = abstract.lower()

    for pattern, design in designs.items():
        if pattern in abstract_lower:
            return design

    return "Not specified"


def _extract_cohort_size(abstract: str) -> str:
    """Extract cohort/sample size from abstract"""
    import re

    # Patterns for sample sizes
    patterns = [
        r'n\s*=\s*(\d+(?:,\d+)?)',
        r'(\d+(?:,\d+)?)\s*patients',
        r'(\d+(?:,\d+)?)\s*participants',
        r'(\d+(?:,\d+)?)\s*subjects',
        r'sample\s*(?:size\s*)?(?:of\s*)?(\d+(?:,\d+)?)',
        r'enrolled\s*(\d+(?:,\d+)?)',
    ]

    for pattern in patterns:
        match = re.search(pattern, abstract.lower())
        if match:
            return match.group(1).replace(',', '')

    return "Not specified"


def _extract_outcomes(abstract: str) -> str:
    """Extract outcome measures from abstract"""
    # Look for outcome-related sentences
    outcome_indicators = [
        'primary outcome', 'secondary outcome', 'endpoint',
        'measured', 'assessed', 'evaluated', 'compared'
    ]

    sentences = abstract.split('. ')

    for sentence in sentences:
        sentence_lower = sentence.lower()
        if any(indicator in sentence_lower for indicator in outcome_indicators):
            # Truncate long sentences
            if len(sentence) > 150:
                return sentence[:150] + '...'
            return sentence

    return "See abstract"


def _extract_key_findings(abstract: str) -> str:
    """Extract key findings/conclusions from abstract"""
    # Look for conclusion indicators
    conclusion_indicators = [
        'conclude', 'conclusion', 'found that', 'results show',
        'demonstrated', 'suggest', 'indicate', 'in summary'
    ]

    sentences = abstract.split('. ')

    # Look from the end (conclusions usually at end)
    for sentence in reversed(sentences):
        sentence_lower = sentence.lower()
        if any(indicator in sentence_lower for indicator in conclusion_indicators):
            if len(sentence) > 200:
                return sentence[:200] + '...'
            return sentence

    # If no conclusion found, take last sentence
    if sentences:
        last = sentences[-1]
        if len(last) > 200:
            return last[:200] + '...'
        return last

    return "See abstract"


def build_literature_matrix(
    papers: List[Dict[str, Any]],
    query: str,
    relevance_scores: Optional[Dict[str, float]] = None,
    fail_closed: bool = True
) -> LiteratureMatrix:
    """
    Build a literature review matrix from papers.

    Args:
        papers: List of paper dictionaries
        query: Research query/topic
        relevance_scores: Optional pre-computed relevance scores by paper ID
        fail_closed: If True, skip papers with PHI

    Returns:
        LiteratureMatrix with extracted information
    """
    rows = []
    relevance_scores = relevance_scores or {}

    for paper in papers:
        try:
            title = paper.get('title', 'Untitled')
            abstract = paper.get('abstract', '') or ''

            # PHI guard
            _, title_findings = guard_text(title, fail_closed=fail_closed)
            _, abstract_findings = guard_text(abstract, fail_closed=fail_closed)

            if (title_findings or abstract_findings) and fail_closed:
                logger.warning(f"PHI detected, skipping paper: {title[:50]}...")
                continue

            # Get identifier
            identifier = (
                paper.get('doi') or
                paper.get('pmid') or
                paper.get('arxivId') or
                paper.get('paperId') or
                'Unknown'
            )

            # Get relevance score
            paper_id = paper.get('pmid') or paper.get('doi') or paper.get('paperId') or title
            relevance = relevance_scores.get(paper_id, 0.5)

            row = MatrixRow(
                title=title,
                year=paper.get('year'),
                identifier=identifier,
                population=_extract_population(abstract),
                design=_extract_study_design(abstract),
                cohort_size=_extract_cohort_size(abstract),
                outcomes=_extract_outcomes(abstract),
                key_findings=_extract_key_findings(abstract),
                notes="",
                relevance_score=relevance
            )

            rows.append(row)

        except Exception as e:
            logger.warning(f"Error processing paper: {e}")
            continue

    # Sort by relevance score
    rows.sort(key=lambda r: r.relevance_score, reverse=True)

    return LiteratureMatrix(
        rows=rows,
        query=query,
        generated_at=datetime.utcnow().isoformat(),
        paper_count=len(rows)
    )


def save_matrix_csv(
    matrix: LiteratureMatrix,
    output_path: str
) -> str:
    """Save matrix as CSV file"""
    import csv

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            'Title', 'Year', 'DOI/PMID', 'Population', 'Design',
            'Cohort Size', 'Outcomes', 'Key Findings', 'Notes', 'Relevance Score'
        ])

        # Data rows
        for row in matrix.rows:
            writer.writerow([
                row.title,
                row.year or '',
                row.identifier,
                row.population,
                row.design,
                row.cohort_size,
                row.outcomes,
                row.key_findings,
                row.notes,
                f"{row.relevance_score:.2f}"
            ])

    return output_path


def save_matrix_xlsx(
    matrix: LiteratureMatrix,
    output_path: str
) -> str:
    """Save matrix as Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Literature Matrix"

        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

        headers = [
            'Title', 'Year', 'DOI/PMID', 'Population', 'Design',
            'Cohort Size', 'Outcomes', 'Key Findings', 'Notes', 'Relevance Score'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row_idx, row in enumerate(matrix.rows, 2):
            ws.cell(row=row_idx, column=1, value=row.title)
            ws.cell(row=row_idx, column=2, value=row.year or '')
            ws.cell(row=row_idx, column=3, value=row.identifier)
            ws.cell(row=row_idx, column=4, value=row.population)
            ws.cell(row=row_idx, column=5, value=row.design)
            ws.cell(row=row_idx, column=6, value=row.cohort_size)
            ws.cell(row=row_idx, column=7, value=row.outcomes)
            ws.cell(row=row_idx, column=8, value=row.key_findings)
            ws.cell(row=row_idx, column=9, value=row.notes)
            ws.cell(row=row_idx, column=10, value=row.relevance_score)

        # Adjust column widths
        ws.column_dimensions['A'].width = 50  # Title
        ws.column_dimensions['B'].width = 8   # Year
        ws.column_dimensions['C'].width = 25  # DOI/PMID
        ws.column_dimensions['D'].width = 30  # Population
        ws.column_dimensions['E'].width = 20  # Design
        ws.column_dimensions['F'].width = 12  # Cohort Size
        ws.column_dimensions['G'].width = 40  # Outcomes
        ws.column_dimensions['H'].width = 50  # Key Findings
        ws.column_dimensions['I'].width = 30  # Notes
        ws.column_dimensions['J'].width = 12  # Relevance

        # Add metadata sheet
        meta_ws = wb.create_sheet("Metadata")
        meta_ws['A1'] = "Query"
        meta_ws['B1'] = matrix.query
        meta_ws['A2'] = "Generated At"
        meta_ws['B2'] = matrix.generated_at
        meta_ws['A3'] = "Paper Count"
        meta_ws['B3'] = matrix.paper_count

        wb.save(output_path)
        return output_path

    except ImportError:
        logger.warning("openpyxl not installed, skipping Excel export")
        return ""


def save_matrix_artifacts(
    matrix: LiteratureMatrix,
    topic_id: str,
    output_dir: Optional[str] = None
) -> LiteratureMatrix:
    """
    Save matrix as both CSV and Excel artifacts.

    Args:
        matrix: LiteratureMatrix to save
        topic_id: Topic/search identifier
        output_dir: Output directory (defaults to ARTIFACTS_PATH)

    Returns:
        Updated LiteratureMatrix with file paths
    """
    base_dir = Path(output_dir or ARTIFACTS_PATH)
    artifact_dir = base_dir / "literature" / topic_id
    artifact_dir.mkdir(parents=True, exist_ok=True)

    # Save CSV
    csv_path = str(artifact_dir / "literature_matrix.csv")
    save_matrix_csv(matrix, csv_path)
    matrix.csv_path = csv_path

    # Save Excel
    xlsx_path = str(artifact_dir / "literature_matrix.xlsx")
    xlsx_result = save_matrix_xlsx(matrix, xlsx_path)
    if xlsx_result:
        matrix.xlsx_path = xlsx_path

    logger.info(f"Saved literature matrix artifacts to {artifact_dir}")
    return matrix
