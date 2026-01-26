"""
Literature Matrix Builder

Builds comparison matrices for literature reviews - author, year, methods, findings.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from src.provenance.artifact_store import store_text, new_run_id

logger = logging.getLogger(__name__)


@dataclass
class LiteratureMatrixConfig:
    """Configuration for literature matrix generation."""
    include_columns: List[str] = field(default_factory=lambda: [
        "title", "authors", "year", "methods", "population",
        "key_findings", "limitations", "venue"
    ])
    output_formats: List[str] = field(default_factory=lambda: ["csv", "markdown"])
    max_cell_length: int = 200
    sort_by: str = "year"  # year, title, author
    sort_descending: bool = True


@dataclass
class MatrixRow:
    """A row in the literature matrix."""
    id: str
    title: str
    authors: str
    year: Optional[int]
    methods: Optional[str]
    population: Optional[str]
    key_findings: List[str]
    limitations: List[str]
    venue: Optional[str]
    doi: Optional[str]
    provider: str


def truncate_text(text: str, max_length: int) -> str:
    """Truncate text to max length with ellipsis."""
    if not text or len(text) <= max_length:
        return text or ""
    return text[:max_length - 3] + "..."


def format_authors(authors: List[Dict[str, Any]]) -> str:
    """Format author list for display."""
    if not authors:
        return ""

    names = []
    for author in authors[:5]:  # Limit to first 5 authors
        if isinstance(author, dict):
            name = author.get("name", "")
        else:
            name = str(author)
        if name:
            names.append(name)

    result = "; ".join(names)
    if len(authors) > 5:
        result += " et al."
    return result


def format_list_for_cell(items: List[str], max_length: int) -> str:
    """Format a list for table cell display."""
    if not items:
        return ""
    text = "; ".join(items)
    return truncate_text(text, max_length)


def build_matrix_rows(
    papers: List[Dict[str, Any]],
    summaries: Optional[List[Dict[str, Any]]] = None,
    config: LiteratureMatrixConfig = None,
) -> List[MatrixRow]:
    """
    Build matrix rows from papers and optional summaries.

    Args:
        papers: List of paper items (LiteratureItem format)
        summaries: Optional list of paper summaries (from summarization job)
        config: Matrix configuration

    Returns:
        List of MatrixRow objects
    """
    if config is None:
        config = LiteratureMatrixConfig()

    # Build summary lookup by paper ID
    summary_lookup = {}
    if summaries:
        for summary in summaries:
            paper_id = summary.get("paper_id", "")
            if paper_id:
                summary_lookup[paper_id] = summary

    rows = []
    for paper in papers:
        paper_id = paper.get("id", "")
        summary = summary_lookup.get(paper_id, {})

        row = MatrixRow(
            id=paper_id,
            title=truncate_text(paper.get("title", ""), config.max_cell_length),
            authors=format_authors(paper.get("authors", [])),
            year=paper.get("year"),
            methods=truncate_text(
                summary.get("methods") or paper.get("methods", ""),
                config.max_cell_length
            ),
            population=truncate_text(
                summary.get("population") or paper.get("population", ""),
                config.max_cell_length
            ),
            key_findings=summary.get("key_findings", []),
            limitations=summary.get("limitations", []),
            venue=paper.get("venue"),
            doi=paper.get("doi"),
            provider=paper.get("provider", "unknown"),
        )
        rows.append(row)

    # Sort rows
    if config.sort_by == "year":
        rows.sort(key=lambda r: r.year or 0, reverse=config.sort_descending)
    elif config.sort_by == "title":
        rows.sort(key=lambda r: r.title.lower(), reverse=config.sort_descending)
    elif config.sort_by == "author":
        rows.sort(key=lambda r: r.authors.lower(), reverse=config.sort_descending)

    return rows


def generate_csv(rows: List[MatrixRow], config: LiteratureMatrixConfig) -> str:
    """Generate CSV output from matrix rows."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    headers = []
    column_map = {
        "title": "Title",
        "authors": "Authors",
        "year": "Year",
        "methods": "Methods",
        "population": "Population",
        "key_findings": "Key Findings",
        "limitations": "Limitations",
        "venue": "Venue/Journal",
        "doi": "DOI",
        "provider": "Source",
    }

    for col in config.include_columns:
        headers.append(column_map.get(col, col.title()))
    writer.writerow(headers)

    # Data rows
    for row in rows:
        row_data = []
        for col in config.include_columns:
            if col == "title":
                row_data.append(row.title)
            elif col == "authors":
                row_data.append(row.authors)
            elif col == "year":
                row_data.append(str(row.year) if row.year else "")
            elif col == "methods":
                row_data.append(row.methods or "")
            elif col == "population":
                row_data.append(row.population or "")
            elif col == "key_findings":
                row_data.append(format_list_for_cell(row.key_findings, config.max_cell_length))
            elif col == "limitations":
                row_data.append(format_list_for_cell(row.limitations, config.max_cell_length))
            elif col == "venue":
                row_data.append(row.venue or "")
            elif col == "doi":
                row_data.append(row.doi or "")
            elif col == "provider":
                row_data.append(row.provider)
        writer.writerow(row_data)

    return output.getvalue()


def generate_markdown(rows: List[MatrixRow], config: LiteratureMatrixConfig) -> str:
    """Generate Markdown table from matrix rows."""
    lines = ["# Literature Matrix", ""]

    # Build header
    headers = []
    column_map = {
        "title": "Title",
        "authors": "Authors",
        "year": "Year",
        "methods": "Methods",
        "population": "Population",
        "key_findings": "Key Findings",
        "limitations": "Limitations",
        "venue": "Venue",
        "doi": "DOI",
        "provider": "Source",
    }

    for col in config.include_columns:
        headers.append(column_map.get(col, col.title()))

    # Header row
    lines.append("| " + " | ".join(headers) + " |")

    # Separator
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

    # Data rows
    for row in rows:
        row_data = []
        for col in config.include_columns:
            value = ""
            if col == "title":
                value = row.title.replace("|", "\\|")
            elif col == "authors":
                value = row.authors.replace("|", "\\|")
            elif col == "year":
                value = str(row.year) if row.year else ""
            elif col == "methods":
                value = (row.methods or "").replace("|", "\\|")
            elif col == "population":
                value = (row.population or "").replace("|", "\\|")
            elif col == "key_findings":
                value = format_list_for_cell(row.key_findings, config.max_cell_length).replace("|", "\\|")
            elif col == "limitations":
                value = format_list_for_cell(row.limitations, config.max_cell_length).replace("|", "\\|")
            elif col == "venue":
                value = (row.venue or "").replace("|", "\\|")
            elif col == "doi":
                value = row.doi or ""
            elif col == "provider":
                value = row.provider
            row_data.append(value)
        lines.append("| " + " | ".join(row_data) + " |")

    return "\n".join(lines)


def generate_xlsx_bytes(rows: List[MatrixRow], config: LiteratureMatrixConfig) -> bytes:
    """Generate XLSX output from matrix rows."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Literature Matrix"

    # Column headers
    column_map = {
        "title": "Title",
        "authors": "Authors",
        "year": "Year",
        "methods": "Methods",
        "population": "Population",
        "key_findings": "Key Findings",
        "limitations": "Limitations",
        "venue": "Venue/Journal",
        "doi": "DOI",
        "provider": "Source",
    }

    headers = [column_map.get(col, col.title()) for col in config.include_columns]

    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        col_idx = 1
        for col in config.include_columns:
            value = ""
            if col == "title":
                value = row.title
            elif col == "authors":
                value = row.authors
            elif col == "year":
                value = row.year
            elif col == "methods":
                value = row.methods or ""
            elif col == "population":
                value = row.population or ""
            elif col == "key_findings":
                value = format_list_for_cell(row.key_findings, config.max_cell_length)
            elif col == "limitations":
                value = format_list_for_cell(row.limitations, config.max_cell_length)
            elif col == "venue":
                value = row.venue or ""
            elif col == "doi":
                value = row.doi or ""
            elif col == "provider":
                value = row.provider

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_idx += 1

    # Auto-adjust column widths
    for col_idx, col in enumerate(config.include_columns, 1):
        max_length = len(column_map.get(col, col))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_literature_matrix(
    papers: List[Dict[str, Any]],
    summaries: Optional[List[Dict[str, Any]]] = None,
    config: Optional[LiteratureMatrixConfig] = None,
    save_artifact: bool = True,
) -> Dict[str, Any]:
    """
    Build a literature matrix from papers and summaries.

    Args:
        papers: List of paper items
        summaries: Optional list of paper summaries
        config: Matrix configuration
        save_artifact: Whether to save as artifact

    Returns:
        Result with matrix data and output files
    """
    if config is None:
        config = LiteratureMatrixConfig()

    if not papers:
        return {
            "paper_count": 0,
            "rows": [],
            "outputs": {},
        }

    logger.info(f"Building literature matrix for {len(papers)} papers")

    # Build rows
    rows = build_matrix_rows(papers, summaries, config)

    # Generate outputs
    outputs = {}
    artifacts = {}

    if "csv" in config.output_formats:
        csv_content = generate_csv(rows, config)
        outputs["csv"] = csv_content

    if "markdown" in config.output_formats:
        md_content = generate_markdown(rows, config)
        outputs["markdown"] = md_content

    if "xlsx" in config.output_formats:
        try:
            xlsx_content = generate_xlsx_bytes(rows, config)
            outputs["xlsx_size"] = len(xlsx_content)
        except Exception as e:
            logger.warning(f"Could not generate XLSX: {e}")

    # Save artifacts
    if save_artifact and outputs:
        try:
            run_id = new_run_id("lit_matrix")

            if "csv" in outputs:
                store_text(
                    run_id=run_id,
                    category="literature_matrix",
                    filename="matrix.csv",
                    text=outputs["csv"],
                )
                artifacts["csv"] = f"{run_id}/matrix.csv"

            if "markdown" in outputs:
                store_text(
                    run_id=run_id,
                    category="literature_matrix",
                    filename="matrix.md",
                    text=outputs["markdown"],
                )
                artifacts["markdown"] = f"{run_id}/matrix.md"

            logger.info(f"Saved literature matrix artifact: {run_id}")

        except Exception as e:
            logger.warning(f"Failed to save artifact: {e}")

    # Convert rows to dicts for JSON serialization
    row_dicts = [
        {
            "id": r.id,
            "title": r.title,
            "authors": r.authors,
            "year": r.year,
            "methods": r.methods,
            "population": r.population,
            "key_findings": r.key_findings,
            "limitations": r.limitations,
            "venue": r.venue,
            "doi": r.doi,
            "provider": r.provider,
        }
        for r in rows
    ]

    return {
        "paper_count": len(papers),
        "rows": row_dicts,
        "outputs": {
            "csv": "csv" in outputs,
            "markdown": "markdown" in outputs,
            "xlsx": "xlsx_size" in outputs,
        },
        "artifacts": artifacts,
        "config": {
            "columns": config.include_columns,
            "sort_by": config.sort_by,
        },
    }
